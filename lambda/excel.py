from openpyxl import Workbook
from openpyxl import load_workbook


#  def checkexcel():
#
#  # 指定したファイルが存在しない場合。
#  FileNotFoundError
#
#  # ファイルにアクセス権がない場合（開いていてロックされている場合など）。
#  PermissionError
#
#  # .xlsx は ZIP 形式なので、壊れている場合や Excel 形式でない場合に発生。
#  zipfile.BadZipFile
#
#  # openpyxl がサポートしていないファイル形式の場合。 .xls .xlsb
#  openpyxl.utils.exceptions.InvalidFileException
#
#  # Excel内部のXMLが壊れている場合。
#  xml.etree.ElementTree.ParseError
#
#  # 内部構造が壊れている Excel ファイルの場合。
#  KeyError




from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
#
# try:
#     wb = load_workbook("sample.xlsx")
# except FileNotFoundError:
#     print("ファイルが存在しません")
# except PermissionError:
#     print("ファイルにアクセスできません")
# except InvalidFileException:
#     print("サポートされていないExcel形式です")
# except BadZipFile:
#     print("壊れたExcelファイルです")
# except Exception as e:
#     print("その他エラー:", e)
#






def openexcel():
    openfile = load_workbook("err.xlsx")

#    openfile = load_workbook("sampleopentest.xlsx")
    ws = openfile.active

    for row in ws.iter_rows(values_only=True):
        print(row)



def readall():
    wb = load_workbook("sample.xlsx")
    ws = wb.active


    data = []
    for row in ws.iter_rows(
            min_row=2,
            # max_row=10,
            min_col=1,
            # max_col=5,
            values_only=True):
        data.append(list(row))

    print(data)



# Excelブック新規作成
def newexcel():
    wb = Workbook()
    ws = wb.active

    # データ書き込み
#    ws["A1"] = "名前"
#    ws["B1"] = "年齢"

    ws.cell(row=1,column=1,value="名前です")
    ws.cell(row=1,column=2,value="年齢です")
    ws.cell(row=1,column=3,value="体重")
    ws.cell(row=1,column=4,value="真偽")
    ws.cell(row=1,column=5,value="誕生日")



    ws.append(["田中", 25, 50, None, "1974/6/24"])
    ws.append(["佐藤", 30, 60.0, None, "1974/6/30"])

#    for i in range(1,2):
#        cell = ws.cell(row=i, column=3)
#        cell.number_format = "0.0000"

    # 保存
    wb.save("sample.xlsx")



# openexcel()
# newexcel()
readall()

